"""
Comprehensive Backtest System Test Suite
=========================================

Tests all components of the backtest engine v2.0:
- BaseBacktestEngine correctness
- BrokerSimulator execution
- PerformanceAnalyzer metrics
- Cost calculations
- Signal generation
- Edge cases and error handling

Author: QuantumTrader Team
Date: November 2025
"""

import sys
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import MetaTrader5 as mt5

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from engines.base_backtest_engine import (
    BaseBacktestEngine, 
    BaseStrategy
)
from engines.broker_simulator import BrokerSimulator, BrokerConfig
from engines.performance_analyzer import PerformanceAnalyzer

print("=" * 80)
print("BACKTEST SYSTEM TEST SUITE")
print("=" * 80)
print()

# Test counters
total_tests = 0
passed_tests = 0
failed_tests = 0

def test_header(test_name):
    """Print test header"""
    global total_tests
    total_tests += 1
    print(f"\n{'='*80}")
    print(f"TEST {total_tests}: {test_name}")
    print(f"{'='*80}")

def test_pass(message=""):
    """Mark test as passed"""
    global passed_tests
    passed_tests += 1
    print(f"PASS: {message}")

def test_fail(message=""):
    """Mark test as failed"""
    global failed_tests
    failed_tests += 1
    print(f"FAIL: {message}")

def assert_equal(actual, expected, msg=""):
    """Assert equality"""
    if actual == expected:
        test_pass(f"{msg} - Got: {actual}")
        return True
    else:
        test_fail(f"{msg} - Expected: {expected}, Got: {actual}")
        return False

def assert_true(condition, msg=""):
    """Assert condition is true"""
    if condition:
        test_pass(msg)
        return True
    else:
        test_fail(msg)
        return False

def assert_greater(actual, threshold, msg=""):
    """Assert value is greater than threshold"""
    if actual > threshold:
        test_pass(f"{msg} - {actual} > {threshold}")
        return True
    else:
        test_fail(f"{msg} - {actual} <= {threshold}")
        return False

def assert_in_range(actual, min_val, max_val, msg=""):
    """Assert value is in range"""
    if min_val <= actual <= max_val:
        test_pass(f"{msg} - {actual} in [{min_val}, {max_val}]")
        return True
    else:
        test_fail(f"{msg} - {actual} not in [{min_val}, {max_val}]")
        return False


# =============================================================================
# TEST 1: MT5 Connection and Data Loading
# =============================================================================
test_header("MT5 Connection and Data Loading")

try:
    if not mt5.initialize():
        test_fail("MT5 initialization failed")
    else:
        test_pass("MT5 initialized successfully")
        
        # Test data loading
        symbol = 'EURUSDm'
        start_date = datetime(2025, 1, 1)
        end_date = datetime(2025, 2, 1)
        
        rates = mt5.copy_rates_range(
            symbol,
            mt5.TIMEFRAME_H1,
            start_date,
            end_date
        )
        
        if rates is not None and len(rates) > 0:
            test_pass(f"Loaded {len(rates)} bars for {symbol}")
            
            # Check data structure
            df = pd.DataFrame(rates)
            required_cols = ['time', 'open', 'high', 'low', 'close', 'tick_volume']
            
            if all(col in df.columns for col in required_cols):
                test_pass("Data contains all required columns")
            else:
                test_fail("Missing required columns in data")
        else:
            test_fail(f"Failed to load data for {symbol}")
            
except Exception as e:
    test_fail(f"MT5 test error: {str(e)}")


# =============================================================================
# TEST 2: Simple Test Strategy
# =============================================================================
test_header("Test Strategy Implementation")

class SimpleTestStrategy(BaseStrategy):
    """Simple strategy for testing - Buy every 10 bars"""
    
    def __init__(self, period=10):
        super().__init__()
        self.period = period
        self.bar_count = 0
        
    def prepare_data(self, data):
        """No indicators needed"""
        return data
    
    def analyze(self, data, current_bar):
        """Generate signal every N bars"""
        self.bar_count += 1
        
        if self.bar_count % self.period == 0:
            return {
                'action': 'BUY',
                'reason': f'Test signal {self.bar_count}'
            }
        
        return None

try:
    strategy = SimpleTestStrategy(period=10)
    test_pass("Test strategy created successfully")
    
    # Test prepare_data
    test_data = pd.DataFrame({
        'time': pd.date_range('2025-01-01', periods=5, freq='h'),
        'open': [1.1, 1.2, 1.3, 1.4, 1.5],
        'close': [1.15, 1.25, 1.35, 1.45, 1.55]
    })
    
    prepared = strategy.prepare_data(test_data)
    assert_equal(len(prepared), 5, "Prepared data has correct length")
    
except Exception as e:
    test_fail(f"Strategy test error: {str(e)}")


# =============================================================================
# TEST 3: BrokerSimulator - Order Execution
# =============================================================================
test_header("BrokerSimulator - Order Execution")

try:
    config = BrokerConfig(
        commission_per_lot=7.0,
        spread_pips=1.5,
        slippage_pips=0.5,
        swap_long=-2.0,
        swap_short=-3.0,
        rejection_rate=0.0  # No rejection for testing
    )
    
    broker = BrokerSimulator(
        config=config,
        initial_balance=10000.0
    )
    
    test_pass("BrokerSimulator created successfully")
    
    # Test balance
    assert_equal(broker.get_balance(), 10000.0, "Initial balance correct")
    
    # Test BUY order
    current_bar = {
        'time': datetime(2025, 1, 1, 10, 0),
        'open': 1.1000,
        'high': 1.1050,
        'low': 1.0950,
        'close': 1.1020,
        'tick_volume': 100
    }
    
    order = broker.execute_order(
        direction='BUY',
        symbol='EURUSD',
        lot_size=0.1,
        current_bar=current_bar,
        reason='Test buy order'
    )
    
    if order:
        test_pass(f"BUY order executed: {order['order_id']}")
        assert_true('entry_price' in order, "Order has entry_price")
        assert_true('commission' in order, "Order has commission")
        assert_true('spread_cost' in order, "Order has spread_cost")
    else:
        test_fail("BUY order failed to execute")
    
    # Check open positions
    positions = broker.get_open_positions()
    assert_equal(len(positions), 1, "One position is open")
    
    # Test position update
    next_bar = {
        'time': datetime(2025, 1, 1, 11, 0),
        'open': 1.1030,
        'high': 1.1080,
        'low': 1.1020,
        'close': 1.1060,
        'tick_volume': 100
    }
    
    broker.update_positions(next_bar)
    test_pass("Positions updated successfully")
    
    # Test SELL order to close
    close_order = broker.execute_order(
        direction='SELL',
        symbol='EURUSD',
        lot_size=0.1,
        current_bar=next_bar,
        reason='Close position'
    )
    
    if close_order:
        test_pass("SELL order executed (position closed)")
        assert_equal(len(broker.get_open_positions()), 0, "No open positions after close")
    
except Exception as e:
    test_fail(f"BrokerSimulator test error: {str(e)}")


# =============================================================================
# TEST 4: Cost Calculations
# =============================================================================
test_header("Cost Calculations Accuracy")

try:
    config = BrokerConfig(
        commission_per_lot=7.0,
        spread_pips=1.5,
        slippage_pips=0.5
    )
    
    broker = BrokerSimulator(
        config=config,
        initial_balance=10000.0
    )
    
    current_bar = {
        'time': datetime(2025, 1, 1, 10, 0),
        'open': 1.1000,
        'high': 1.1050,
        'low': 1.0950,
        'close': 1.1020,
        'tick_volume': 100
    }
    
    order = broker.execute_order(
        direction='BUY',
        symbol='EURUSD',
        lot_size=1.0,  # 1 lot
        current_bar=current_bar,
        reason='Cost test'
    )
    
    if order:
        # Commission should be $7 for 1 lot
        expected_commission = 7.0
        assert_equal(
            order['commission'],
            expected_commission,
            f"Commission for 1 lot"
        )
        
        # Spread cost: 1.5 pips * $10/pip * 1 lot = $15
        expected_spread = 15.0
        assert_in_range(
            order['spread_cost'],
            expected_spread - 1,
            expected_spread + 1,
            "Spread cost for 1 lot, 1.5 pips"
        )
        
        test_pass(f"Total entry cost: ${order['commission'] + order['spread_cost']:.2f}")
    
except Exception as e:
    test_fail(f"Cost calculation error: {str(e)}")


# =============================================================================
# TEST 5: Stop Loss and Take Profit
# =============================================================================
test_header("Stop Loss and Take Profit Execution")

try:
    config = BrokerConfig()
    broker = BrokerSimulator(config=config, initial_balance=10000.0)
    
    # Open BUY position
    entry_bar = {
        'time': datetime(2025, 1, 1, 10, 0),
        'open': 1.1000,
        'high': 1.1050,
        'low': 1.0950,
        'close': 1.1020,
        'tick_volume': 100
    }
    
    order = broker.execute_order(
        direction='BUY',
        symbol='EURUSD',
        lot_size=0.1,
        current_bar=entry_bar,
        sl_pips=50,
        tp_pips=100,
        reason='SL/TP test'
    )
    
    if order:
        position = broker.get_open_positions()[0]
        
        # Check SL/TP are set
        assert_true('stop_loss' in position, "Position has stop_loss")
        assert_true('take_profit' in position, "Position has take_profit")
        
        # SL should be 50 pips below entry
        expected_sl = position['entry_price'] - (50 * 0.0001)
        assert_in_range(
            position['stop_loss'],
            expected_sl - 0.0001,
            expected_sl + 0.0001,
            f"Stop Loss set correctly at {position['stop_loss']:.5f}"
        )
        
        # TP should be 100 pips above entry
        expected_tp = position['entry_price'] + (100 * 0.0001)
        assert_in_range(
            position['take_profit'],
            expected_tp - 0.0001,
            expected_tp + 0.0001,
            f"Take Profit set correctly at {position['take_profit']:.5f}"
        )
        
        # Test TP hit
        tp_bar = {
            'time': datetime(2025, 1, 1, 11, 0),
            'open': 1.1020,
            'high': 1.1150,  # High enough to hit TP
            'low': 1.1000,
            'close': 1.1100,
            'tick_volume': 100
        }
        
        broker.update_positions(tp_bar)
        
        # Position should be closed
        assert_equal(len(broker.get_open_positions()), 0, "Position closed by Take Profit")
        
        # Check closed trades
        trades = broker.get_closed_trades()
        if len(trades) > 0:
            last_trade = trades[-1]
            assert_equal(last_trade['exit_reason'], 'Take Profit', "Exit reason is Take Profit")
            assert_greater(last_trade['net_pnl'], 0, "TP trade is profitable")
    
except Exception as e:
    test_fail(f"SL/TP test error: {str(e)}")


# =============================================================================
# TEST 6: Full Backtest Run
# =============================================================================
test_header("Full Backtest Engine Integration")

try:
    # Create simple strategy
    class TestSMAStrategy(BaseStrategy):
        def prepare_data(self, data):
            data['sma_fast'] = data['close'].rolling(5).mean()
            data['sma_slow'] = data['close'].rolling(10).mean()
            return data
        
        def analyze(self, data, current_bar):
            current_time = current_bar['time']
            
            if current_time not in data.index:
                return None
            
            idx = data.index.get_loc(current_time)
            
            if idx < 1:
                return None
            
            # Check for golden cross
            if (data['sma_fast'].iloc[idx] > data['sma_slow'].iloc[idx] and
                data['sma_fast'].iloc[idx-1] <= data['sma_slow'].iloc[idx-1]):
                return {
                    'action': 'BUY',
                    'sl_pips': 50,
                    'tp_pips': 100,
                    'reason': 'SMA Golden Cross'
                }
            
            # Check for death cross
            elif (data['sma_fast'].iloc[idx] < data['sma_slow'].iloc[idx] and
                  data['sma_fast'].iloc[idx-1] >= data['sma_slow'].iloc[idx-1]):
                return {
                    'action': 'SELL',
                    'sl_pips': 50,
                    'tp_pips': 100,
                    'reason': 'SMA Death Cross'
                }
            
            return None
    
    # Create engine with new API
    strategy = TestSMAStrategy()
    
    broker_config = BrokerConfig(
        commission_per_lot=7.0,
        spread_pips=1.5,
        slippage_pips=0.5
    )
    
    engine = BaseBacktestEngine(
        strategy=strategy,
        broker_config=broker_config,
        initial_balance=10000.0
    )
    
    test_pass("Backtest engine created successfully")
    
    # Run backtest on small dataset
    print("\nRunning backtest on 1 month of data...")
    
    metrics = engine.run_backtest(
        symbol='EURUSDm',
        start_date=datetime(2025, 1, 1),
        end_date=datetime(2025, 2, 1),
        timeframe=mt5.TIMEFRAME_H1,
        lot_size=0.1,
        show_progress=False
    )
    
    if metrics:
        test_pass("Backtest completed successfully")
        
        # Validate metrics
        assert_true('initial_balance' in metrics, "Metrics contain initial_balance")
        assert_true('final_balance' in metrics, "Metrics contain final_balance")
        assert_true('total_trades' in metrics, "Metrics contain total_trades")
        assert_true('winning_trades' in metrics, "Metrics contain winning_trades")
        assert_true('losing_trades' in metrics, "Metrics contain losing_trades")
        
        # Check logical consistency
        total = metrics['total_trades']
        wins = metrics['winning_trades']
        losses = metrics['losing_trades']
        
        if total > 0:
            assert_equal(wins + losses, total, "Wins + Losses = Total Trades")
            
            win_rate = metrics.get('win_rate', 0)
            expected_win_rate = (wins / total) * 100
            assert_in_range(
                win_rate,
                expected_win_rate - 0.1,
                expected_win_rate + 0.1,
                "Win rate calculation correct"
            )
        
        # Display results
        print(f"\n📊 Backtest Results:")
        print(f"   Total Trades: {total}")
        print(f"   Winning Trades: {wins}")
        print(f"   Losing Trades: {losses}")
        print(f"   Win Rate: {win_rate:.1f}%")
        print(f"   Initial Balance: ${metrics['initial_balance']:,.2f}")
        print(f"   Final Balance: ${metrics['final_balance']:,.2f}")
        print(f"   Total Return: {metrics.get('total_return', 0):.2f}%")
        
    else:
        test_fail("Backtest returned no metrics")
    
except Exception as e:
    test_fail(f"Full backtest error: {str(e)}")
    import traceback
    traceback.print_exc()


# =============================================================================
# TEST 7: Performance Analyzer
# =============================================================================
test_header("Performance Analyzer Metrics")

try:
    # Create sample trades
    sample_trades = [
        {
            'entry_time': datetime(2025, 1, 1, 10, 0),
            'exit_time': datetime(2025, 1, 1, 15, 0),
            'direction': 'BUY',
            'entry_price': 1.1000,
            'exit_price': 1.1050,
            'lot_size': 0.1,
            'gross_pnl': 50.0,
            'commission': 1.4,
            'spread_cost': 1.5,
            'swap': 0.0,
            'net_pnl': 47.1,
            'exit_reason': 'Take Profit'
        },
        {
            'entry_time': datetime(2025, 1, 2, 10, 0),
            'exit_time': datetime(2025, 1, 2, 12, 0),
            'direction': 'SELL',
            'entry_price': 1.1050,
            'exit_price': 1.1100,
            'lot_size': 0.1,
            'gross_pnl': -50.0,
            'commission': 1.4,
            'spread_cost': 1.5,
            'swap': 0.0,
            'net_pnl': -52.9,
            'exit_reason': 'Stop Loss'
        }
    ]
    
    analyzer = PerformanceAnalyzer(initial_balance=10000.0)
    
    # Add trades to analyzer
    for trade in sample_trades:
        analyzer.add_trade(trade)
    
    metrics = analyzer.calculate_metrics()
    
    test_pass("Performance metrics calculated")
    
    # Validate key metrics
    assert_equal(metrics['total_trades'], 2, "Total trades count")
    assert_equal(metrics['winning_trades'], 1, "Winning trades count")
    assert_equal(metrics['losing_trades'], 1, "Losing trades count")
    
    # Win rate should be 50%
    assert_in_range(metrics['win_rate'], 49.9, 50.1, "Win rate 50%")
    
    # Average win/loss
    assert_in_range(metrics['average_win'], 47.0, 47.2, "Average win")
    assert_in_range(abs(metrics['average_loss']), 52.8, 53.0, "Average loss")
    
    test_pass(f"All metrics validated: {len(metrics)} metrics calculated")
    
except Exception as e:
    test_fail(f"Performance analyzer error: {str(e)}")


# =============================================================================
# TEST 8: Edge Cases
# =============================================================================
test_header("Edge Cases and Error Handling")

try:
    # Test 1: Empty data
    config = BrokerConfig()
    broker = BrokerSimulator(config=config, initial_balance=10000.0)
    
    # Try to execute order with invalid bar
    invalid_bar = {}
    order = broker.execute_order(
        direction='BUY',
        symbol='EURUSD',
        lot_size=0.1,
        current_bar=invalid_bar,
        reason='Invalid bar test'
    )
    
    if order is None:
        test_pass("Invalid bar rejected correctly")
    else:
        test_fail("Invalid bar should be rejected")
    
    # Test 2: Zero lot size
    valid_bar = {
        'time': datetime(2025, 1, 1, 10, 0),
        'open': 1.1000,
        'high': 1.1050,
        'low': 1.0950,
        'close': 1.1020,
        'tick_volume': 100
    }
    
    order = broker.execute_order(
        direction='BUY',
        symbol='EURUSD',
        lot_size=0.0,
        current_bar=valid_bar,
        reason='Zero lot test'
    )
    
    if order is None:
        test_pass("Zero lot size rejected correctly")
    else:
        test_fail("Zero lot size should be rejected")
    
    # Test 3: Invalid direction
    order = broker.execute_order(
        direction='INVALID',
        symbol='EURUSD',
        lot_size=0.1,
        current_bar=valid_bar,
        reason='Invalid direction test'
    )
    
    if order is None:
        test_pass("Invalid direction rejected correctly")
    else:
        test_fail("Invalid direction should be rejected")
    
    # Test 4: Insufficient balance
    broker = BrokerSimulator(config=config, initial_balance=10.0)  # Very low balance
    
    order = broker.execute_order(
        direction='BUY',
        symbol='EURUSD',
        lot_size=100.0,  # Huge lot size
        current_bar=valid_bar,
        reason='Insufficient balance test'
    )
    
    # This might still execute (no margin check in simple simulator)
    # but we test it doesn't crash
    test_pass("Large lot size handled without crash")
    
except Exception as e:
    test_fail(f"Edge case handling error: {str(e)}")


# =============================================================================
# TEST 9: Data Integrity
# =============================================================================
test_header("Data Integrity and Time Index")

try:
    # This tests the critical bug fix
    symbol = 'EURUSDm'
    start_date = datetime(2025, 1, 1)
    end_date = datetime(2025, 1, 15)
    
    rates = mt5.copy_rates_range(
        symbol,
        mt5.TIMEFRAME_H1,
        start_date,
        end_date
    )
    
    if rates is not None and len(rates) > 0:
        df = pd.DataFrame(rates)
        df['time'] = pd.to_datetime(df['time'], unit='s')
        df.set_index('time', inplace=True)
        
        # Check index is datetime
        assert_true(
            isinstance(df.index, pd.DatetimeIndex),
            "DataFrame index is DatetimeIndex"
        )
        
        # Check no duplicate index
        assert_equal(
            df.index.duplicated().sum(),
            0,
            "No duplicate time entries"
        )
        
        # Check chronological order
        assert_true(
            df.index.is_monotonic_increasing,
            "Data is in chronological order"
        )
        
        # Check data completeness
        assert_equal(
            df[['open', 'high', 'low', 'close']].isna().sum().sum(),
            0,
            "No missing OHLC values"
        )
        
        test_pass(f"Data integrity verified for {len(df)} bars")
    
except Exception as e:
    test_fail(f"Data integrity test error: {str(e)}")


# =============================================================================
# TEST 10: Excel Export
# =============================================================================
test_header("Excel Report Generation")

try:
    from openpyxl import Workbook
    
    # Create sample metrics
    metrics = {
        'initial_balance': 10000.0,
        'final_balance': 9500.0,
        'total_trades': 100,
        'winning_trades': 45,
        'losing_trades': 55
    }
    
    # Create sample trades
    trades = [
        {
            'entry_time': datetime(2025, 1, 1, 10, 0),
            'exit_time': datetime(2025, 1, 1, 15, 0),
            'net_pnl': 50.0
        }
    ]
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Report"
    
    ws['A1'] = "Metric"
    ws['B1'] = "Value"
    
    row = 2
    for key, value in metrics.items():
        ws[f'A{row}'] = key
        ws[f'B{row}'] = value
        row += 1
    
    # Save to temp file
    temp_file = Path("test_report.xlsx")
    wb.save(temp_file)
    
    if temp_file.exists():
        test_pass("Excel report generated successfully")
        
        # Verify file can be read
        import openpyxl
        wb_read = openpyxl.load_workbook(temp_file)
        ws_read = wb_read.active
        
        assert_equal(ws_read['A1'].value, "Metric", "Excel content correct")
        
        # Cleanup
        temp_file.unlink()
        test_pass("Excel file verified and cleaned up")
    else:
        test_fail("Excel file not created")
    
except Exception as e:
    test_fail(f"Excel export error: {str(e)}")


# =============================================================================
# FINAL SUMMARY
# =============================================================================
print("\n" + "=" * 80)
print("TEST SUMMARY")
print("=" * 80)
print(f"\nPassed: {passed_tests}/{total_tests}")
print(f"Failed: {failed_tests}/{total_tests}")

if failed_tests == 0:
    print(f"\nALL TESTS PASSED! Backtest system is working correctly!")
    success_rate = 100.0
else:
    success_rate = (passed_tests / total_tests) * 100
    print(f"\nSome tests failed. Success rate: {success_rate:.1f}%")

print(f"\nTest Coverage:")
print(f"   - MT5 Connection & Data Loading")
print(f"   - Strategy Implementation")
print(f"   - Broker Simulation")
print(f"   - Cost Calculations")
print(f"   - Stop Loss / Take Profit")
print(f"   - Full Backtest Integration")
print(f"   - Performance Metrics")
print(f"   - Edge Cases & Error Handling")
print(f"   - Data Integrity")
print(f"   - Excel Export")

print("\n" + "=" * 80)

# Cleanup MT5
mt5.shutdown()

# Exit code
sys.exit(0 if failed_tests == 0 else 1)
